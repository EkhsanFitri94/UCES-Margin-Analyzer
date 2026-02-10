import streamlit as st
import pandas as pd
import io
import os
import json
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Protection, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Configuration ---
st.set_page_config(page_title="UCES Margin Analyzer", layout="wide")

# --- CSS Styling ---
st.markdown("""
<style>
    .stApp { background-color: #f5f5f5; }
    h1 { color: #2e4053; }
    .info-box {
        background-color: #e2e3e5;
        border: 1px solid #d6d8db;
        padding: 10px;
        border-radius: 5px;
        margin-bottom: 15px;
        font-size: 0.9em;
    }
    div[data-testid="stExpander"] {
        background-color: white;
        border-radius: 10px;
        border: 1px solid #ddd;
        border-left: 5px solid #2e4053;
    }
    .filter-container {
        background-color: white;
        padding: 15px;
        border-radius: 10px;
        border: 1px solid #ddd;
        margin-bottom: 20px;
    }
    .legend-box {
        position: sticky;
        top: 0;
        z-index: 999;
        background-color: white;
        padding: 10px;
        border-radius: 5px;
        border: 1px solid #ddd;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
    }
</style>
""", unsafe_allow_html=True)

# --- PERSISTENCE FUNCTIONS (JSON VERSION) ---
DATA_FILE = os.path.join(os.getcwd(), "uces_app_data.json")

def save_data():
    """Saves current state to JSON file."""
    df_to_save = st.session_state.df.copy()
    # Convert date objects to string for JSON serialization
    if 'Date of PR' in df_to_save.columns:
        df_to_save['Date of PR'] = df_to_save['Date of PR'].astype(str)
    # Convert Margin Reason to string
    if 'Margin Reason' in df_to_save.columns:
        df_to_save['Margin Reason'] = df_to_save['Margin Reason'].astype(str)
        
    data_to_save = {
        'df': df_to_save.to_dict(orient='records'),
        'source_filename': st.session_state.source_filename
    }
    try:
        with open(DATA_FILE, "w") as f:
            json.dump(data_to_save, f)
    except Exception as e:
        pass 

def load_data():
    """Loads data from JSON file if it exists."""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r") as f:
                data = json.load(f)
                df = pd.DataFrame(data['df'])
                return {
                    'df': df,
                    'source_filename': data['source_filename']
                }
        except Exception:
            return None
    return None

# --- Helper Function: Initialize DF ---
def init_df():
    return pd.DataFrame(columns=[
        "Quotation No", "Po Huawei", "Linked PR Subcon", "Date of PR", "Vendor Name", "Project", "Site ID", "Line Items",
        "Po Huawei (Unit Price)", "Requested Qty", "Total", 
        "Po Subcon (Unit Price)", "Qty", "Sub Total", "Profit", "Margin%", "Status", "Margin Reason"
    ])

# --- Initialize Session State ---
if 'df' not in st.session_state:
    saved_data = load_data()
    if saved_data:
        st.session_state.df = saved_data['df']
        st.session_state.source_filename = saved_data['source_filename']
    else:
        st.session_state.df = init_df()
        st.session_state.source_filename = "master_file_data.xlsx"

# Define Project Options
PROJECT_OPTIONS = {
    "---": "(NON-PROJECT)",
    "BD": "Business Development",
    "CME": "Civil Mechanical Electrical",
    "CS": "Customer Support",
    "HQ": "Head Quarter",
    "IBS": "Inbuilding System",
    "MISC": "Miscellaneous Project",
    "MS": "Managing Services",
    "RNO": "Radio Network Optimization",
    "SOLAR": "Solar",
    "TI": "Technical Installation",
    "TINSOL": "Tinno Solar"
}
PROJECT_KEYS = list(PROJECT_OPTIONS.keys())

# --- App Title ---
st.title("📊 UCES Margin Analyzer")
st.markdown("Check Client PO vs Subcon PO Margins (Target: ≥30%)")

# --- STICKY LEGEND (EXPLANATION OF COLORS) ---
with st.container():
    st.markdown('<div class="legend-box">', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("### 🟢 ≥ 30% (Healthy)")
    with c2:
        st.markdown("### 🟡 20–29% (Below Target)")
    with c3:
        st.markdown("### 🔴 < 20% (Loss Risk)")
    st.markdown('</div>', unsafe_allow_html=True)

# --- Section 1: Upload Excel ---
with st.expander("📥 Step 1: Load Existing Excel File", expanded=False):
    uploaded_file = st.file_uploader("Choose your Master Excel file", type=['xlsx', 'xls'], key="file_uploader")
    
    if uploaded_file is not None:
        st.info(f"File selected: `{uploaded_file.name}`. Click Confirm to Load.")
        if st.button("🔄 Load File", type="primary", key="load_btn"):
            try:
                temp_df = pd.read_excel(uploaded_file)
                required = init_df().columns.tolist()
                uploaded_cols = temp_df.columns.tolist()
                
                required_lower = [x.lower() for x in required]
                uploaded_cols_lower = [x.lower() for x in uploaded_cols]
                cols_match = all(item in uploaded_cols_lower for item in required_lower)
                
                if cols_match and len(temp_df) > 0:
                    col_map = {x.lower(): x for x in required}
                    temp_df.columns = [col_map.get(str(c).lower(), c) for c in temp_df.columns]
                    
                    # FORCE TYPE CASTING: Ensure 'Date of PR' is datetime
                    if 'Date of PR' in temp_df.columns:
                        temp_df['Date of PR'] = pd.to_datetime(temp_df['Date of PR'], errors='coerce')
                    
                    # Fill NaN with empty strings to prevent crashes
                    temp_df = temp_df.fillna("")
                    
                    st.session_state.df = temp_df[required]
                    st.session_state.source_filename = uploaded_file.name
                    save_data() 
                    st.success(f"Loaded {len(temp_df)} rows from {uploaded_file.name}!")
                    st.rerun()
                else:
                    st.markdown(f"<div class='info-box'>⚠️ <b>Column Mismatch.</b> Please map your columns.</div>", unsafe_allow_html=True)
                    
                    mapping = {}
                    cols_with_none = ["(Ignore/Missing)"] + uploaded_cols
                    st.write("Map columns:")
                    map_cols = st.columns(2)
                    
                    for i, col in enumerate(required):
                        default_idx = 0
                        simple_col = col.replace(" ", "").replace("(", "").replace(")", "").lower()
                        for j, up_col in enumerate(uploaded_cols):
                            simple_up = up_col.replace(" ", "").replace("(", "").replace(")", "").lower()
                            if simple_col == simple_up:
                                default_idx = j + 1
                                break
                        
                        with map_cols[i % 2]:
                            selected = st.selectbox(f"System: `{col}`", cols_with_none, index=default_idx, key=f"map_{col}")
                            mapping[col] = selected if selected != "(Ignore/Missing)" else None
                    
                    if st.button("Apply Mapping & Load", type="secondary", key="map_load_btn"):
                        final_df = pd.DataFrame()
                        for app_col, excel_col in mapping.items():
                            if excel_col is not None:
                                final_df[app_col] = temp_df[excel_col]
                            else:
                                if app_col == "Status": final_df[app_col] = "Process"
                                elif app_col == "Project": final_df[app_col] = "---"
                                elif app_col == "Margin Status": final_df[app_col] = ""
                                elif app_col == "Margin Reason": final_df[app_col] = ""
                                elif app_col == "Date of PR": final_df[app_col] = ""
                                elif app_col in ["Po Huawei (Unit Price)", "Po Subcon (Unit Price)", "Total", "Sub Total", "Profit", "Margin%"]: final_df[app_col] = 0.0
                                elif app_col in ["Requested Qty", "Qty"]: final_df[app_col] = 0
                                else: final_df[app_col] = ""
                        st.session_state.df = final_df
                        st.session_state.source_filename = uploaded_file.name
                        save_data() 
                        st.success("Loaded mapped data!")
                        st.rerun()

            except Exception as e:
                st.error(f"Error: {e}")

st.divider()

# --- Section 2: Form (Add/Edit) ---
if 'edit_index' not in st.session_state:
    st.session_state.edit_index = None

form_mode = "✏️ Edit Entry" if st.session_state.edit_index is not None else "➕ Add New Entry"

with st.expander(f"{form_mode}", expanded=(st.session_state.edit_index is not None)):
    
    if st.session_state.edit_index is not None:
        # ATOMIC FIX: Deep copy to ensure independence
        row_data = st.session_state.df.iloc[st.session_state.edit_index].copy(deep=True)
        
        current_project = row_data.get("Project", "---")
        if current_project != "---" and current_project in PROJECT_OPTIONS:
            proj_default = current_project
        else:
            proj_default = "---" 

        # SAFE DATE RETRIEVAL
        raw_date_val = row_data.get("Date of PR")
        safe_date_val = date.today()
        if pd.notna(raw_date_val):
            if isinstance(raw_date_val, str):
                try:
                    parsed = pd.to_datetime(raw_date_val, errors='coerce', dayfirst=True)
                    if pd.notna(parsed):
                        safe_date_val = parsed.date()
                except:
                    pass
            elif hasattr(raw_date_val, 'date'):
                try:
                    safe_date_val = raw_date_val.date()
                except:
                    safe_date_val = raw_date_val
            elif isinstance(raw_date_val, date):
                safe_date_val = raw_date_val

        # ROBUST DEFAULTS DICT (Always define all keys)
        default_vals = {
            "quotation_no": str(row_data.get("Quotation No", "")),
            "po_huawei": str(row_data.get("Po Huawei", "")),
            "linked_pr": str(row_data.get("Linked PR Subcon", "")),
            "date_pr": safe_date_val,
            "vendor_name": str(row_data.get("Vendor Name", "")),
            "project": proj_default,
            "site_id": str(row_data.get("Site ID", "")),
            "line_items": str(row_data.get("Line Items", "")),
            "margin_reason": str(row_data.get("Margin Reason", "")), # Safe .get() for Edit mode
            "status": row_data.get("Status", "Waiting"),
            "hp": float(row_data.get("Po Huawei (Unit Price)", 0.0)),
            "rq": int(row_data.get("Requested Qty", 0)),
            "sp": float(row_data.get("Po Subcon (Unit Price)", 0.0)),
            "sq": int(row_data.get("Qty", 0)),
        }
        btn_text = "💾 Update Row"
        btn_type = "primary"
    else:
        # ADD MODE: Defaults
        default_vals = {
            "quotation_no": "",
            "po_huawei": "", "linked_pr": "", "date_pr": date.today(),
            "vendor_name": "", "project": "---", "site_id": "", "line_items": "", "margin_reason": "", # Empty string for Add mode
            "status": "Waiting",
            "hp": 0.0, "rq": 0, "sp": 0.0, "sq": 0
        }
        btn_text = "➕ Add to Table"
        btn_type = "secondary"

    with st.form("data_form", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            quotation_no = st.text_input("Quotation No", value=default_vals["quotation_no"])
            
            po_huawei = st.text_input("PO Huawei*", value=default_vals["po_huawei"])
            linked_pr = st.text_input("Linked PR Subcon", value=default_vals["linked_pr"])
            
            date_pr = st.date_input("Date of PR", value=default_vals["date_pr"], format="DD/MM/YYYY")
        
        with col2:
            vendor_name = st.text_input("Vendor Name", value=default_vals["vendor_name"])
            
            site_id = st.text_input("Site ID", value=default_vals["site_id"])
            
            status = st.selectbox("Status", ["Waiting", "Process", "Rejected", "Claimed"], 
                                  index=["Waiting", "Process", "Rejected", "Claimed"].index(default_vals["status"]) if default_vals["status"] in ["Waiting", "Process", "Rejected", "Claimed"] else 0)
        
        with col3:
            st.subheader("Financials")
            
            project = st.selectbox("Project", PROJECT_KEYS, 
                                  index=PROJECT_KEYS.index(default_vals["project"]) if default_vals["project"] in PROJECT_KEYS else 0,
                                  format_func=lambda x: f"{x} = {PROJECT_OPTIONS[x]}")
            
            final_line_item = st.text_input("Line Items", value=default_vals["line_items"], placeholder="e.g. Router X5")
            
            po_huawei_price = st.number_input("PO Huawei (Unit Price)", value=default_vals["hp"], format="%.2f")
            req_qty = st.number_input("Requested Qty", value=default_vals["rq"], step=1)
            po_subcon_price = st.number_input("PO Subcon (Unit Price)", value=default_vals["sp"], format="%.2f")
            subcon_qty = st.number_input("Qty (Subcon)", value=default_vals["sq"], step=1)

            # MOVED HERE: Margin Reason Text Area (Visible in Financials section)
            margin_reason = st.text_area("Margin Reason", value=default_vals["margin_reason"], placeholder="e.g. Vendor raised price...", height=70)

        submitted = st.form_submit_button(btn_text, type=btn_type)
        cancel_edit = st.form_submit_button("Cancel Edit") if st.session_state.edit_index is not None else None

        if submitted:
            if not po_huawei:
                st.error("PO Huawei is required!")
            else:
                total_huawei = po_huawei_price * req_qty
                total_subcon = po_subcon_price * subcon_qty
                profit = total_huawei - total_subcon
                margin = (profit / total_huawei * 100) if total_huawei != 0 else 0.0

                # --- AUTO-FLAG LOGIC ---
                if margin >= 30.0:
                    margin_status = "Healthy"
                elif margin >= 20.0:
                    margin_status = "Below Target"
                else:
                    margin_status = "Loss Risk"

                # Use margin_status for both add and edit
                final_status_text = margin_reason.strip() if margin_reason.strip() != "" else margin_status

                new_row = {
                    "Quotation No": quotation_no,
                    "Po Huawei": po_huawei,
                    "Linked PR Subcon": linked_pr,
                    "Date of PR": date_pr,
                    "Vendor Name": vendor_name,
                    "Project": project,
                    "Site ID": site_id,
                    "Line Items": final_line_item,
                    "Po Huawei (Unit Price)": po_huawei_price,
                    "Requested Qty": req_qty,
                    "Total": total_huawei,
                    "Po Subcon (Unit Price)": po_subcon_price,
                    "Qty": subcon_qty,
                    "Sub Total": total_subcon,
                    "Profit": profit,
                    "Margin%": round(margin, 2),
                    "Status": final_status_text,
                    "Margin Reason": margin_reason
                }

                if st.session_state.edit_index is not None:
                    st.session_state.df.iloc[st.session_state.edit_index] = new_row
                    st.session_state.edit_index = None
                    st.success("Row updated successfully!")
                else:
                    st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
                    st.success("Entry added successfully!")
                save_data()
                st.rerun()
        
        if cancel_edit:
            st.session_state.edit_index = None
            st.rerun()

# --- Section 3: View Data (WITH FILTERS) ---
st.divider()
st.subheader("Current Master File Data")

# --- FILTERS SECTION ---
with st.container():
    st.markdown('<div class="filter-container">', unsafe_allow_html=True)
    
    # Row 1: Status
    filter_status = st.selectbox("🔍 Filter by Status", ["All", "Waiting", "Process", "Rejected", "Claimed"])
    
    # Row 2: Text Filters (Project, Vendor, Site ID)
    col_2a, col_2b, col_2c = st.columns(3)
    with col_2a:
        search_project = st.text_input("🔍 Project Code", placeholder="e.g. BD, SOLAR")
    with col_2b:
        search_vendor = st.text_input("🔍 Vendor Name", placeholder="Vendor Name")
    with col_2c:
        search_site = st.text_input("🔍 Site ID", placeholder="Site ID")
    
    # Row 3: Margin Filter
    margin_filter = st.selectbox("🔍 Filter by Margin", ["All", "Loss Risk (<20%)", "Below Target (20-29%)", "Healthy (≥30%)"])
    
    # Row 4: Reset Button
    if st.button("🔄 Reset All Filters"):
        st.rerun()
            
    st.markdown('</div>', unsafe_allow_html=True)

if not st.session_state.df.empty:
    # FILTER LOGIC
    filtered_df = st.session_state.df.copy()

    if filter_status != "All":
        filtered_df = filtered_df[filtered_df["Status"] == filter_status]

    if search_project and not filtered_df.empty:
        filtered_df = filtered_df[filtered_df["Project"].str.contains(search_project, case=False, na=False)]

    if search_vendor and not filtered_df.empty:
        filtered_df = filtered_df[filtered_df["Vendor Name"].str.contains(search_vendor, case=False, na=False)]

    if search_site and not filtered_df.empty:
        filtered_df = filtered_df[filtered_df["Site ID"].str.contains(search_site, case=False, na=False)]

    if margin_filter == "Loss Risk (<20%)":
        filtered_df = filtered_df[filtered_df["Margin%"] < 20.0]
    elif margin_filter == "Below Target (20-29%)":
        filtered_df = filtered_df[(filtered_df["Margin%"] >= 20.0) & (filtered_df["Margin%"] < 30.0)]
    elif margin_filter == "Healthy (≥30%)":
        filtered_df = filtered_df[filtered_df["Margin%"] >= 30.0]

    # --- FIX: Ensure 'Date of PR' is always pandas datetime64 for Arrow compatibility ---
    if "Date of PR" in filtered_df.columns:
        filtered_df["Date of PR"] = pd.to_datetime(filtered_df["Date of PR"], errors="coerce")

    st.caption(f"Showing {len(filtered_df)} of {len(st.session_state.df)} records")

    # TABLE STYLING
    def color_margin(val):
        try:
            val_float = float(val)
            if val_float >= 30.0:
                return 'background-color: #d1fae5; color: green; font-weight: bold' # Healthy (Green)
            elif val_float >= 20.0:
                return 'background-color: #fff3cd; color: #856404; font-weight: bold' # Warning (Yellow-Orange)
            else:
                return 'background-color: #f8d7da; color: white; font-weight: bold' # Risk (Red background, white text)
        except:
            return ''

    def color_text(val):
        # Color text itself (Dark Green vs Dark Red) based on margin
        try:
            val_float = float(val)
            if val_float >= 30.0:
                return 'color: #006400' # Dark Green
            else:
                return 'color: #dc2626' # Dark Red
        except:
            return 'color: black'

    # Use .map instead of deprecated .applymap
    styled_df = filtered_df.style.map(color_margin, subset=['Margin%']).map(color_text, subset=['Margin%'])

    column_config = {
        "Date of PR": st.column_config.DateColumn("Date", format="DD/MM/YYYY"),
        "Po Huawei (Unit Price)": st.column_config.NumberColumn("Price (RM)", format="%.2f"),
        "Requested Qty": st.column_config.NumberColumn("Req Qty", format="%d"), 
        "Total": st.column_config.NumberColumn("Total (RM)", format="%.2f"),
        "Po Subcon (Unit Price)": st.column_config.NumberColumn("Sub Price (RM)", format="%.2f"),
        "Qty": st.column_config.NumberColumn("Sub Qty", format="%d"), 
        "Sub Total": st.column_config.NumberColumn("Sub Total (RM)", format="%.2f"),
        "Profit": st.column_config.NumberColumn("Profit (RM)", format="%.2f"),
        "Margin%": st.column_config.NumberColumn("Margin%", format="%.2f"),
        "Status": st.column_config.TextColumn("Status", width="medium"),
        "Margin Reason": st.column_config.TextColumn("Reason", width="large"),
        "Project": st.column_config.SelectboxColumn(
            "Project",
            options=PROJECT_KEYS,
            required=True
        )
    }

    st.dataframe(
        styled_df,
        use_container_width=True,
        height=500,
        column_config=column_config
    )

    st.markdown("---")
    st.markdown("### Quick Actions")
    action_col1, action_col2, action_col3 = st.columns([2, 1, 1])
    
    with action_col1:
        options = [f"Row {i}: {r['Po Huawei']} - {r['Site ID']}" for i, r in st.session_state.df.iterrows()]
        selected_row_opt = st.selectbox("Select a row to manage:", options, index=None)
        
    with action_col2:
        if selected_row_opt:
            selected_index = int(selected_row_opt.split(":")[0].split(" ")[1])
            if st.button("✏️ Edit Selected", type="primary", use_container_width=True):
                st.session_state.edit_index = selected_index
                st.rerun()
                
    with action_col3:
        if selected_row_opt:
            if st.button("🗑️ Delete Selected", type="secondary", use_container_width=True):
                st.session_state.df = st.session_state.df.drop(selected_index)
                st.session_state.df = st.session_state.df.reset_index(drop=True)
                save_data() 
                st.rerun()

else:
    st.info("No data to display. Upload a file or add a new entry.")

# --- Section 4: Footer / Download ---
st.divider()

st.markdown("### 💾 Save Your Work")
st.info("Download your updated Excel file here.")

col_dl, col_clr = st.columns(2)

with col_dl:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        st.session_state.df.to_excel(writer, index=False, sheet_name='Master File')
        
        workbook = writer.book
        worksheet = workbook['Master File']
        
        green_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        yellow_orange_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        col_indices = {}
        for col in range(1, worksheet.max_column + 1):
            header_val = worksheet.cell(row=1, column=col).value
            if header_val:
                col_indices[header_val] = col
        
        margin_col_idx = col_indices.get("Margin%")
        status_col_idx = col_indices.get("Status")
        reason_col_idx = col_indices.get("Margin Reason")
        
        project_col_letter = None
        for k, v in col_indices.items():
            if k == "Project":
                project_col_letter = get_column_letter(v)
                break
        
        # Excel Validation
        if project_col_letter and worksheet.max_row > 1:
            dv_range = f"{project_col_letter}2:{project_col_letter}{worksheet.max_row}"
            
            dv = DataValidation(type="list", formula1=f'"{",".join(PROJECT_KEYS)}"', allow_blank=True)
            
            dv.error = 'Your entry is not in list'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = 'Please select from the list'
            dv.promptTitle = 'Project Selection'
            
            dv.add(dv_range)
            worksheet.add_data_validation(dv)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.protection = Protection(locked=False)
                
                cell_header = worksheet.cell(row=1, column=cell.column).value
                
                if cell_header in ["Po Huawei (Unit Price)", "Total", "Po Subcon (Unit Price)", "Sub Total", "Profit"]:
                    cell.number_format = '#,##0.00'
                elif cell_header in ["Requested Qty", "Qty"]:
                    cell.number_format = '0'
                elif cell_header == "Margin%":
                    cell.number_format = '0.00'
                    try:
                        val = float(cell.value)
                        if val >= 30.0:
                            cell.fill = green_fill
                            cell.font = Font(color="006400", bold=True)
                        elif val >= 20.0:
                            cell.fill = yellow_orange_fill
                            cell.font = Font(color="856404", bold=True)
                        else:
                            cell.fill = red_fill
                            cell.font = Font(color="FFFFFF", bold=True) # White text on red
                    except:
                        pass
                elif cell_header == "Date of PR":
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.number_format = 'General'

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    buffer.seek(0)
    
    st.download_button(
        label=f"📥 Update & Download File ({st.session_state.source_filename})",
        data=buffer,
        file_name=st.session_state.source_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with col_clr:
    if st.button("🗑️ Clear All Data", type="primary"):
        st.session_state.df = init_df()
        st.session_state.edit_index = None
        st.rerun()
